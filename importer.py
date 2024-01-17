from openpyxl import load_workbook
import random
import json
from datetime import datetime as dt

file_name = input("Enter file name: ")
if not file_name:
    print("No file specified.")
    exit()

t_amount_missing = []
wb = load_workbook(file_name)
ws = wb.active

headings = []
for col in range(1, ws.max_column + 1):
    headings.append(ws.cell(row=1, column=col).value)

#print(headings)
known_r_no = []
p_size_missing = []
booking_date_missing = []

with open("data.json", "r+") as f:
    data = json.load(f)

colony_name = input("Enter colony name: ")
if colony_name not in data:
    data[colony_name] = {}
    data[colony_name]["sectors"] ={}

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):    
    s_no = row[0].value
    if not s_no:
        continue
    try:
        if row[1].font.color.rgb == "FFFF0000":
            continue
    except AttributeError as e:
        pass
    p_no = row[1].value
    party = row[2].value
    if not party:
        continue
    address = row[3].value
    mobile_no = row[4].value
    if not row[5].value:
        p_size_missing.append(p_no)
        continue
    p_size = row[5].value.lower()
    char = "*" if "*" in p_size else "x"
    p_size = [float(x) for x in p_size.split(char)]
    area = row[6].value
    if not area:
        area = p_size[0]*p_size[1]
    advisor = row[9].value
    if not advisor:
        continue
    booking_date = row[10].value

    reciepts = []
    total = 0
    col = 8
    while True:
        col += 3
        if headings[col] == "R.TOTAL":
            break
        cr_no = row[col].value
        if isinstance(cr_no, str):
            if "/" in cr_no:
                cr_no = cr_no.split("/")[0]
            if "\\" in cr_no:
                cr_no = cr_no.split("\\")[0] 
            if "," in cr_no:
                cr_no = cr_no.split(",")[0]
            if " " in cr_no:
                cr_no = cr_no.split(" ")[0]
            cr_no = int(cr_no)
        
        known_r_no.append(cr_no)
        date = row[col + 1].value
        amount = row[col + 2].value
        if not amount:
            continue
        if not date:
            date = headings[col+2]
        #print(cr_no, date, amount)
        total = total+amount
        r = {
            "reciept_number": 0,
            "date": "",
            "amount": 0,
            "mode": "cash",
            "is_cheque": False
            }
        r["reciept_number"] = cr_no
        if isinstance(date, str):
            date_str = date
            try:
                date = dt.strptime(date_str , "%d-%m-%Y")
            except ValueError as v:
                try:
                    date = date_str.split("/")[0]
                    date = dt.strptime(date, "%d-%m-%Y")
                except ValueError:
                    date = date_str.split("/")[-1] # last one 
                    date = dt.strptime(date, "%d-%m-%Y")
        if not booking_date:
            booking_date = date
        r["date"] = date.strftime("%Y-%m-%d")
        r["amount"] = amount
        reciepts.append(r)
    r_total = row[col].value
    t_amount = row[col+1].value
    remark = row[col+3].value
    status = "booked"
    if remark:
        if "ok" in str(remark).lower():
            status = "registered"
    if len(reciepts) == 1 and t_amount == reciepts[0]["amount"]:
        emi=False
    else:
        emi=True
    if not total == r_total:
        print("Total mismatch")
        print(total, r_total, p_no)
        break
    if not t_amount:
        t_amount_missing.append(p_no)
        continue
    template = {
        "size": "",
        "rate": "",
        "price": "",
        "deal_price": "",
        "booking_amount": "",
        "incentive": "",
        "customer": {},
        "status": "available",
        "is_emi": "",
        "reciept_entry": "",
        "advisor": "",
        "date": "",
        "files": []
        }
    rate = t_amount/area 
    # limit rate to 2 decimal places
    rate = int(rate*100)/100
    template["size"] = p_size
    template["rate"] = rate
    template["price"] = t_amount
    template["deal_price"] = t_amount
    template["booking_amount"] = reciepts[0]["amount"] if len(reciepts) > 0 else 0
    template["incentive"] = 10
    template["customer"]["name"] = party
    template["customer"]["address"] = address
    template["customer"]["mobile_no"] = mobile_no
    template["status"] = status
    template["is_emi"] = emi
    template["reciept_entry"] = reciepts
    template["advisor"] = advisor.lower().strip().capitalize()
    if not booking_date:
        booking_date_missing.append(p_no)
        continue
    template["date"] = booking_date.strftime("%Y-%m-%d")
    template["files"] = []
    sector = p_no.split("-")
    pno = sector[1]
    if pno.startswith("0"):
        pno = pno[1:]
    sector = sector[0].upper()
    if sector not in data[colony_name]["sectors"]:
        data[colony_name]["sectors"][sector] = {}
        data[colony_name]["sectors"][sector]["plots"] = {}
    if pno in data[colony_name]["sectors"][sector]["plots"]:
        for x in range(1,11):
            if pno+"-"+str(x) not in data[colony_name]["sectors"][sector]["plots"]:
                pno = pno+"-"+str(x)
                break
    data[colony_name]["sectors"][sector]["plots"][pno] = template

known_cr_no = {}
for colony in data:
    known_cr_no[colony] = []
    for sector in data[colony]["sectors"]:
        for plot in data[colony]["sectors"][sector]["plots"]:
            for reciept in data[colony]["sectors"][sector]["plots"][plot]["reciept_entry"]:
                if reciept["reciept_number"] is not None:
                    known_cr_no[colony].append(int(reciept["reciept_number"]))

for colony in data:
    maximum_known_cr_no = max(known_cr_no[colony]) + 151
    for sector in data[colony]["sectors"]:
        for plot in data[colony]["sectors"][sector]["plots"]:
            for reciept in data[colony]["sectors"][sector]["plots"][plot]["reciept_entry"]:
                if reciept["reciept_number"] is None:
                    maximum_known_cr_no += 1
                    reciept["reciept_number"] = maximum_known_cr_no
                    print(f"{sector}-{plot} {reciept['date']}", maximum_known_cr_no)
                    known_cr_no[colony].append(maximum_known_cr_no)
"""
missing = []
for x in range(min(known_cr_no), max(known_cr_no)):
    if x not in known_cr_no:
        missing.append(x)
print(f"Missing - {' '.join([str(x) for x in missing])}")"""

with open("data.json", "w+") as f:
    f.write(json.dumps(data, indent=4))
print("total missing: ")
print(" ".join(t_amount_missing))
print("plot size missing: ")
print(" ".join(p_size_missing))
print("booking date missing: ")
print(" ".join(booking_date_missing))
